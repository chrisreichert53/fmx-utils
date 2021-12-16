import os, requests, json
from requests_toolbelt import threaded
from pandas import DataFrame
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from itertools import islice
from pydash import invert, count_by, find, get, last
import json_flatten as flat

import numpy as np
from pandas import DataFrame

load_dotenv()  # take environment variables from .env.

from util import changes, status_report
from util.questions import do_add, which_buildings, which_fields


URL = os.getenv("URL")
FMX_USERNAME = os.getenv("FMX_USERNAME")
FMX_PASSWORD = os.getenv("FMX_PASSWORD")
FILE_PATH = os.getenv("FILE_PATH")
SHEET_BUILDING_FIELD = os.getenv("SHEET_BUILDING_FIELD")
SHEET_TAG_FIELD = os.getenv("SHEET_TAG_FIELD")
SHEET_TYPE_FIELD = os.getenv("SHEET_TYPE_FIELD")
SHEET_ID_FIELD = os.getenv("SHEET_ID_FIELD")
SHEET_LOCATION_FIELD = os.getenv("SHEET_LOCATION_FIELD")
NUMBER_FIELDS = os.getenv("NUMBER_FIELDS")


def initialize_session(session):
    session.auth = (FMX_USERNAME, FMX_PASSWORD)
    session.headers["Content-Type"] = "application/json"
    return session


s = requests.Session()
initialize_session(s)

wb = load_workbook(filename=FILE_PATH, data_only=True)
sheet: Worksheet = wb[wb.sheetnames[0]]

data = sheet.values
cols = next(data)[0:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 0, None) for r in data)
df = DataFrame(data, index=idx, columns=cols)
contents = df.to_dict("records")

site_equipment_options = s.get(f"https://{URL}/api/v1/equipment/get-options").json()
site_building_ids_org = site_equipment_options["buildings"]
site_building_ids = invert(site_building_ids_org)
equipment_type_ids = invert(site_equipment_options["equipmentTypes"])
site_location_ids = invert(site_equipment_options["resources"])
custom_field_ids = invert({**site_equipment_options["sortKeys"], **site_equipment_options["customFields"]})
# print(json.dumps(custom_field_ids, indent=2))

# Which buildings to include in update?
sheet_buildings = df[SHEET_BUILDING_FIELD].unique()
sheet_buildings = np.sort(sheet_buildings)
sheet_buildings = [
    {"name": bldg, "id": site_building_ids[bldg]} for bldg in sheet_buildings if bldg in site_building_ids
]
counts = count_by(contents, lambda obj: obj[SHEET_BUILDING_FIELD])
update_buildings = which_buildings(
    [
        {"name": bldg["name"] + " (" + str(counts[bldg["name"]]) + ")", "value": bldg}
        for bldg in sheet_buildings
    ]
)
buildings = (
    sheet_buildings
    if find(update_buildings, lambda bldg: bldg["name"] == "all") is not None
    else update_buildings
)

# Which equipment to include?
filter_buildings = [bldg["name"] for bldg in buildings]
to_be_changed = df[df[SHEET_BUILDING_FIELD].isin(filter_buildings)]

building_equipment = []
for index, row in to_be_changed.iterrows():
    new_piece = {
        "id": row[SHEET_ID_FIELD],
        "tag": row[SHEET_TAG_FIELD],
        "buildingID": site_building_ids[row[SHEET_BUILDING_FIELD]],
        "equipmentTypeID": equipment_type_ids[row[SHEET_TYPE_FIELD]],
    }
    if row[SHEET_LOCATION_FIELD] is not None and row[SHEET_BUILDING_FIELD] is not None:
        new_piece.update(
            {
                "locationResourceID": site_location_ids[
                    row[SHEET_LOCATION_FIELD] + " (" + row[SHEET_BUILDING_FIELD] + ")"
                ]
            }
        )
    new_piece.update(
        {
            "customFields": [
                {
                    "customFieldID": custom_field_ids[id],
                    "value": str(row[id]) if id not in NUMBER_FIELDS else row[id],
                    "name": id,
                }
                for id in custom_field_ids.keys()
                if id in row
            ]
        }
    )
    building_equipment.append(flat.flatten(new_piece))

# Which fields to include?
fields = which_fields(building_equipment)
changeset = changes(building_equipment, fields)

for change in changeset:
    if "customFields" in change:
        change["customFields"] = [
            cf for cf in change["customFields"] if cf["value"] is not None and cf["value"] != "None"
        ]

# Create and pool the requests
reqs = [
    {
        "method": "PUT",
        "url": f"https://{URL}/api/v1/equipment/" + str(change["id"]),
        "data": json.dumps(change),
    }
    for change in changeset
    if get(change, "id") is not None
]
# print(json.dumps(reqs, indent=2))

# If building has some sort of changeset, run the following
results = {"successes": [], "failures": [], "success": 0, "fail": 0, "not_created": 0}
responses, exceptions = (
    threaded.map(reqs, initializer=initialize_session, num_processes=10) if reqs != [] else ([], [])
)
status_report(results, responses, to_be_changed.shape[0], reqs)

# Ask to generate new equipment
if results["not_created"] > 0:
    if do_add(results["not_created"]):
        # Which fields to include?
        fields = which_fields(building_equipment)
        changeset = changes(building_equipment, fields)
        reqs = [
            {
                "method": "POST",
                "url": f"https://{URL}/api/v1/equipment",
                "data": json.dumps(change),
            }
            for change in changeset
            if get(change, "id") is None
        ]
        results = {"successes": [], "failures": [], "success": 0, "fail": 0, "not_created": 0}
        responses, exceptions = threaded.map(reqs, initializer=initialize_session, num_processes=10)
        responses = status_report(results, responses, len(reqs), reqs)

        for res in responses:
            id = res.get("id")
            for row in sheet.iter_rows():
                if row[3].value == site_building_ids_org[str(res["buildingID"])] + res["tag"]:
                    row[2].value = str(id)

        wb.save(FILE_PATH)
